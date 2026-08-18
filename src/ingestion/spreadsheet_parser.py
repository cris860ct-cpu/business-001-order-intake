from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook

from .extraction import ParsedDocument, SourceBlock, normalize_whitespace


HEADER_SYNONYMS = {
    'line_number': {'line_number', 'line', 'line no', 'line #'},
    'customer_item_number': {'customer_item_number', 'customer item number', 'item', 'customer item'},
    'description': {'description', 'item description', 'raw_description'},
    'quantity': {'quantity', 'qty', 'ordered qty'},
    'unit': {'unit', 'uom'},
    'unit_price': {'unit_price', 'price', 'unit price'},
}

META_HINTS = {'customer', 'bill to', 'po', 'po number', 'purchase order', 'order date', 'requested delivery'}


def _cell_text(cell) -> str:
    return normalize_whitespace('' if cell.value is None else str(cell.value))


def _normalized(value: str) -> str:
    return normalize_whitespace(value).lower().replace('#', '').replace('-', ' ').replace('_', ' ')


def _metadata_lines(values: List[str]) -> List[str]:
    pairs: List[str] = []
    for idx, value in enumerate(values):
        if not value:
            continue
        low = value.lower()
        if any(h in low for h in META_HINTS):
            if ':' in value:
                pairs.append(value)
            elif idx + 1 < len(values) and values[idx + 1]:
                pairs.append(f'{value}: {values[idx + 1]}')
    return pairs


def _find_header_row(ws) -> Optional[int]:
    for row_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
        values = [_cell_text(cell) for cell in row]
        normalized = [_normalized(v) for v in values if v]
        hits = 0
        for v in normalized:
            for canon, aliases in HEADER_SYNONYMS.items():
                if any(alias in v for alias in aliases):
                    hits += 1
                    break
        if hits >= 3:
            return row_idx
    return None


def _header_map(values: List[str]) -> Dict[str, int]:
    mapping: Dict[str, int] = {}
    for idx, value in enumerate(values):
        nv = _normalized(value)
        if not nv:
            continue
        for canon, aliases in HEADER_SYNONYMS.items():
            if any(alias in nv for alias in aliases):
                mapping.setdefault(canon, idx)
    return mapping


def _standard_line_text(values: List[str], header_map: Dict[str, int]) -> Optional[str]:
    def get(canon: str) -> str:
        idx = header_map.get(canon)
        if idx is None or idx >= len(values):
            return ''
        return values[idx]

    line_number = get('line_number')
    customer_item = get('customer_item_number')
    description = get('description')
    quantity = get('quantity')
    unit = get('unit')
    price = get('unit_price')
    if not any([line_number, customer_item, description, quantity, unit, price]):
        return None
    if not line_number:
        return None
    return ' | '.join([line_number, customer_item, description, quantity, unit, price])


def parse_workbook_bytes(data: bytes, source_name: str) -> ParsedDocument:
    wb = load_workbook(BytesIO(data), data_only=True)
    return _parse_workbook(wb, source_name)


def parse_workbook(path: str | Path) -> ParsedDocument:
    p = Path(path)
    wb = load_workbook(p, data_only=True)
    return _parse_workbook(wb, str(p))


def _parse_workbook(wb, source_path: str) -> ParsedDocument:
    blocks: List[SourceBlock] = []
    for ws in wb.worksheets[:1]:
        header_row_idx = _find_header_row(ws) or 1

        for row_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
            if row_idx >= header_row_idx:
                break
            values = [_cell_text(cell) for cell in row]
            for line in _metadata_lines(values):
                blocks.append(SourceBlock(source_ref=f'{ws.title}:row{row_idx}', text=line, source_type='xlsx'))

        header_values = [_cell_text(cell) for cell in next(ws.iter_rows(min_row=header_row_idx, max_row=header_row_idx, values_only=False))]
        header_map = _header_map(header_values)
        blocks.append(SourceBlock(source_ref=f'{ws.title}:row{header_row_idx}', text=' | '.join(header_values), source_type='xlsx'))

        for row_idx, row in enumerate(ws.iter_rows(min_row=header_row_idx + 1, values_only=False), start=header_row_idx + 1):
            values = [_cell_text(cell) for cell in row]
            if not any(values):
                continue
            if any(v.lower().startswith('total') for v in values if v):
                blocks.append(SourceBlock(source_ref=f'{ws.title}:row{row_idx}', text=' | '.join(values), source_type='xlsx'))
                continue
            line_text = _standard_line_text(values, header_map)
            if line_text:
                blocks.append(SourceBlock(source_ref=f'{ws.title}:row{row_idx}', text=line_text, source_type='xlsx'))
    raw_text = '\n'.join(block.text for block in blocks)
    return ParsedDocument(source_path=source_path, source_type='xlsx', blocks=blocks, raw_text=raw_text)
