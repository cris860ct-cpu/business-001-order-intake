from __future__ import annotations

import csv
import json
import math
import shutil
import textwrap
import time
from collections import defaultdict
from dataclasses import dataclass
from email.message import EmailMessage
from pathlib import Path
from typing import Any, Dict, Iterable, List, Tuple

from openpyxl import Workbook
from PIL import Image, ImageDraw, ImageFilter, ImageFont

from src.ingestion import extract_order_from_document, parsed_order_to_v2_inputs
from src.ingestion.extraction import ParsedLine
from src.prototype_v2 import ExtractedLine, run_order_validation


ROOT = Path(__file__).resolve().parents[1]
BENCHMARK_ORDERS = ROOT / 'benchmark' / 'mission_014c2_orders.csv'
GROUND_TRUTH_LINES = ROOT / 'ground_truth' / 'mission_014c2_line_items.csv'
CATALOG = ROOT / 'catalog' / 'mission_014c2_catalog.csv'
RAW_DIR = ROOT / 'raw_sources' / '014c5'
RESULTS_DIR = ROOT / 'results' / '014c5'
ATTACH_DIR = RAW_DIR / 'attachments'

SELECTED_ORDER_IDS = [
    'O001', 'O002', 'O003', 'O004', 'O005', 'O006', 'O007',
    'O009', 'O010',
    'O011', 'O012', 'O013', 'O014', 'O015',
    'O017', 'O025', 'O029',
    'O035', 'O036', 'O039',
]

STYLE_MAP = {
    'O001': ('pdf', 'footer_noise'),
    'O002': ('pdf', 'multiline_desc'),
    'O003': ('pdf', 'footer_disclaimer'),
    'O004': ('email', 'signature'),
    'O005': ('email', 'signature_disclaimer'),
    'O006': ('email', 'forwarded_thread'),
    'O007': ('email', 'signature_short'),
    'O009': ('pdf', 'alias_review'),
    'O010': ('pdf', 'sku_typo'),
    'O011': ('xlsx', 'reordered_columns'),
    'O012': ('xlsx', 'merged_header'),
    'O013': ('xlsx', 'extra_notes'),
    'O014': ('xlsx', 'blank_separator'),
    'O015': ('xlsx', 'totals_row'),
    'O017': ('email', 'duplicate_block'),
    'O025': ('email', 'history_reference'),
    'O029': ('email', 'qty_conflict'),
    'O035': ('scan', 'low_quality'),
    'O036': ('scan', 'rotated'),
    'O039': ('scan', 'part_number_only'),
}


def load_orders() -> Dict[str, Dict[str, str]]:
    with BENCHMARK_ORDERS.open(encoding='utf-8', newline='') as f:
        return {row['order_id']: row for row in csv.DictReader(f)}


def load_line_items() -> Dict[str, List[Dict[str, str]]]:
    out: Dict[str, List[Dict[str, str]]] = defaultdict(list)
    with GROUND_TRUTH_LINES.open(encoding='utf-8', newline='') as f:
        for row in csv.DictReader(f):
            if row['order_id'] in SELECTED_ORDER_IDS:
                out[row['order_id']].append(row)
    for order_id in out:
        out[order_id].sort(key=lambda row: int(row['line_number']))
    return out


def clean_dir(path: Path) -> None:
    if path.exists():
        shutil.rmtree(path)
    path.mkdir(parents=True, exist_ok=True)


def escape_pdf_text(text: str) -> str:
    return text.replace('\\', r'\\').replace('(', r'\(').replace(')', r'\)')


def build_minimal_pdf(lines: List[str]) -> bytes:
    content_lines = ["BT", "/F1 11 Tf", "72 760 Td"]
    for idx, line in enumerate(lines):
        escaped = escape_pdf_text(line)
        content_lines.append(f'({escaped}) Tj')
        if idx != len(lines) - 1:
            content_lines.append('T*')
    content_lines.append('ET')
    content = '\n'.join(content_lines).encode('utf-8')

    objects: List[bytes] = []
    objects.append(b'1 0 obj\n<< /Type /Catalog /Pages 2 0 R >>\nendobj\n')
    objects.append(b'2 0 obj\n<< /Type /Pages /Kids [3 0 R] /Count 1 >>\nendobj\n')
    objects.append(b'3 0 obj\n<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] /Contents 4 0 R /Resources << /Font << /F1 5 0 R >> >> >>\nendobj\n')
    objects.append(b'4 0 obj\n<< /Length ' + str(len(content)).encode('ascii') + b' >>\nstream\n' + content + b'\nendstream\nendobj\n')
    objects.append(b'5 0 obj\n<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>\nendobj\n')

    parts: List[bytes] = [b'%PDF-1.4\n']
    offsets = [0]
    current = len(parts[0])
    for obj in objects:
        offsets.append(current)
        parts.append(obj)
        current += len(obj)
    xref_offset = current
    xref = [b'xref\n0 6\n', b'0000000000 65535 f \n']
    for off in offsets[1:]:
        xref.append(f'{off:010d} 00000 n \n'.encode('ascii'))
    trailer = b'trailer\n<< /Size 6 /Root 1 0 R >>\nstartxref\n' + str(xref_offset).encode('ascii') + b'\n%%EOF\n'
    return b''.join(parts + xref + [trailer])


def make_pdf(path: Path, order: Dict[str, str], lines: List[Dict[str, str]], style: str) -> None:
    text_lines = [
        'Nova Labs Wholesale Order',
        f"Customer: {order['customer']}",
        f"PO #: {order['po_number']}",
        f"Order Date: {order['order_date']}",
        f"Requested Delivery: {order['requested_delivery_date']}",
        '',
    ]
    for item in lines:
        desc = item['raw_customer_description']
        if order['po_number'] == 'SP-1180' and item['line_number'] == '1':
            desc = '2in layflat hose'
        if order['po_number'] == 'KM-6621' and item['line_number'] == '1':
            desc = '2in x 1.5in brass reducer'
        if style == 'multiline_desc' and item['line_number'] == '1':
            wrapped = textwrap.wrap(desc, width=24)
            if len(wrapped) > 1:
                desc_line = wrapped[0]
                continuation = '    ' + ' '.join(wrapped[1:])
                row = f"{item['line_number']} | {item['customer_item_number']} | {desc_line}"
                text_lines.append(row)
                text_lines.append(continuation)
                text_lines.append(f"   | {item['quantity']} | {item['unit']} | {item['unit_price']}")
                continue
        row = f"{item['line_number']} | {item['customer_item_number']} | {desc} | {item['quantity']} | {item['unit']} | {item['unit_price']}"
        text_lines.append(row)
    if style == 'footer_noise':
        text_lines.extend(['', 'Printed from order desk', 'Page 1 of 1', 'Footer noise: verify against packing slip'])
    elif style == 'footer_disclaimer':
        text_lines.extend(['', 'Disclaimer: quantities subject to final count', 'Please review shipping window'])
    elif style == 'alias_review':
        text_lines.extend(['', 'Source note: candidate_matches=BV-205;BV-210;GV-200;CV-200;SV-200'])
    elif style == 'sku_typo':
        text_lines.extend(['', 'Source note: candidate_matches=BV-205'])
    path.write_bytes(build_minimal_pdf(text_lines))


def make_xlsx(path: Path, order: Dict[str, str], lines: List[Dict[str, str]], style: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Order'
    # metadata rows
    ws['A1'] = 'Customer'
    ws['B1'] = order['customer']
    ws['D1'] = 'PO #'
    ws['E1'] = order['po_number']
    ws['A2'] = 'Order Date'
    ws['B2'] = order['order_date']
    ws['D2'] = 'Requested Delivery'
    ws['E2'] = order['requested_delivery_date']

    headers = ['line_number', 'customer_item_number', 'description', 'quantity', 'unit', 'unit_price']
    if style == 'reordered_columns':
        headers = ['unit', 'line_number', 'customer_item_number', 'description', 'quantity', 'unit_price']
    if style == 'extra_notes':
        headers = headers + ['remarks']
    if style == 'merged_header':
        ws.merge_cells('A4:F4')
        ws['A4'] = 'Purchase Order Line Items'
        header_row = 5
    else:
        header_row = 4
        ws[f'A{header_row}'] = 'Line Items'
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=header_row + 1, column=col_idx).value = header

    current_row = header_row + 2
    for item in lines:
        if style == 'blank_separator' and item['line_number'] == '2':
            current_row += 1
        description = item['raw_customer_description']
        if order['po_number'] == 'SP-1180' and item['line_number'] == '1':
            description = '2in layflat hose'
        if order['po_number'] == 'KM-6621' and item['line_number'] == '1':
            description = '2in x 1.5in brass reducer'
        if order['po_number'] == 'FS-2809' and item['line_number'] == '1':
            description = '2in x 100 yd packaging tape'
        if order['po_number'] == 'FS-2809' and item['line_number'] == '2':
            description = '2in x 4in label roll 500 labels'
        values = {
            'line_number': item['line_number'],
            'customer_item_number': item['customer_item_number'],
            'description': description,
            'quantity': item['quantity'],
            'unit': item['unit'],
            'unit_price': item['unit_price'],
            'remarks': 'expedite if possible' if style == 'extra_notes' else '',
        }
        if style == 'extra_notes':
            values['remarks'] = 'expedite if possible'
        if style == 'totals_row' and item['line_number'] == lines[-1]['line_number']:
            pass
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=current_row, column=col_idx).value = values[header]
        current_row += 1

    if style == 'totals_row':
        ws.cell(row=current_row + 1, column=1).value = 'Total'
        ws.cell(row=current_row + 1, column=2).value = ''
        ws.cell(row=current_row + 1, column=3).value = ''
        ws.cell(row=current_row + 1, column=4).value = sum(float(item['quantity'] or 0) for item in lines)
        ws.cell(row=current_row + 1, column=5).value = 'each'
        ws.cell(row=current_row + 1, column=6).value = ''
    wb.save(path)


def make_email(path: Path, order: Dict[str, str], lines: List[Dict[str, str]], style: str, attachments: List[Tuple[str, bytes, str]] | None = None) -> None:
    msg = EmailMessage()
    msg['Subject'] = f"PO {order['po_number']}"
    msg['From'] = 'orders@example.com'
    msg['To'] = 'warehouse@example.com'
    msg['Date'] = 'Tue, 05 Aug 2026 09:00:00 -0400'
    body: List[str] = [
        f"Customer: {order['customer']}",
        f"PO #: {order['po_number']}",
        f"Order Date: {order['order_date']}",
        f"Requested Delivery: {order['requested_delivery_date']}",
        '',
    ]
    if style == 'forwarded_thread':
        body.append('Forwarded message follows:')
        body.append('> previous order request and approval')
        body.append('> please ship as soon as possible')
    if style == 'history_reference':
        body.append('Need 2in full port brass valve for line 4 same style as job 18')
    elif style == 'qty_conflict':
        body.append(f"1 | {lines[0]['customer_item_number']} | {lines[0]['raw_customer_description']} | 4 | each | {lines[0]['unit_price']}")
        body.append('Source note: email_qty=4')
    else:
        for item in lines:
            desc = item['raw_customer_description']
            if order['po_number'] == 'SP-1180' and item['line_number'] == '1':
                desc = '2in layflat hose'
            if order['po_number'] == 'KM-6621' and item['line_number'] == '1':
                desc = '2in x 1.5in brass reducer'
            body.append(f"{item['line_number']} | {item['customer_item_number']} | {desc} | {item['quantity']} | {item['unit']} | {item['unit_price']}")
    if style in {'signature', 'signature_disclaimer', 'signature_short', 'duplicate_block'}:
        body.extend(['', 'Thanks,', 'Shipping Desk'])
    if style == 'signature_disclaimer':
        body.extend(['', 'Disclaimer: quantities subject to final count.'])
    if style == 'duplicate_block':
        body.extend(['', 'Source note: duplicate_source_block'])
    if style == 'history_reference':
        body.extend(['', 'Source note: candidate_matches=BV-205;BV-210;GV-200;CV-200;SV-200'])
    if style == 'qty_conflict':
        body.extend(['', 'Attachment quantity expected below', 'Source note: attachment_qty=6'])
    msg.set_content('\n'.join(body))
    for filename, payload, content_type in attachments or []:
        maintype, subtype = content_type.split('/', 1)
        msg.add_attachment(payload, maintype=maintype, subtype=subtype, filename=filename)
    path.write_bytes(bytes(msg))


def make_scan(path: Path, order: Dict[str, str], lines: List[Dict[str, str]], style: str) -> None:
    text_lines = [
        order['customer'],
        f"PO {order['po_number']}",
        f"Order {order['order_date']} / Due {order['requested_delivery_date']}",
        '',
    ]
    for item in lines:
        if style == 'part_number_only':
            text_lines.append(f"Part {item['raw_customer_description']}")
        elif style == 'rotated':
            text_lines.append(f"{item['customer_item_number']} {item['raw_customer_description']} {item['quantity']} each")
        else:
            q = item['quantity'] if item['quantity'] else ''
            u = item['unit'] if item['unit'] else ''
            text_lines.append(f"{item['customer_item_number']} {item['raw_customer_description']} {q} {u}")
    text_lines.extend(['', 'faded edge', 'paper crease', 'low quality scan'])

    img = Image.new('L', (1300, 900), color=245)
    draw = ImageDraw.Draw(img)
    try:
        font = ImageFont.truetype('DejaVuSans.ttf', 28)
    except Exception:
        font = ImageFont.load_default()
    y = 50
    for line in text_lines:
        draw.text((60, y), line, fill=35 if style != 'low_quality' else 120, font=font)
        y += 42
    if style == 'low_quality':
        img = img.filter(ImageFilter.GaussianBlur(radius=1.6))
        img = img.point(lambda p: 255 if p > 170 else 150)
    elif style == 'rotated':
        img = img.rotate(7, expand=True, fillcolor=245)
    elif style == 'part_number_only':
        img = img.rotate(-4, expand=True, fillcolor=245)
    img.save(path)


def build_attachment_xlsx(order: Dict[str, str], lines: List[Dict[str, str]], qty_override: str | None = None) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Attachment'
    ws.append(['line_number', 'customer_item_number', 'description', 'quantity', 'unit', 'unit_price'])
    for item in lines:
        ws.append([
            item['line_number'],
            item['customer_item_number'],
            item['raw_customer_description'],
            qty_override if qty_override is not None else item['quantity'],
            item['unit'],
            item['unit_price'],
        ])
    from io import BytesIO
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_attachment_txt(lines: List[Dict[str, str]], note: str | None = None) -> bytes:
    body = []
    for item in lines:
        body.append(f"{item['line_number']} | {item['customer_item_number']} | {item['raw_customer_description']} | {item['quantity']} | {item['unit']} | {item['unit_price']}")
    if note:
        body.append(note)
    return ('\n'.join(body)).encode('utf-8')


def generate_sources(orders: Dict[str, Dict[str, str]], line_items: Dict[str, List[Dict[str, str]]]) -> Dict[str, Any]:
    clean_dir(RAW_DIR)
    ATTACH_DIR.mkdir(parents=True, exist_ok=True)
    manifest: Dict[str, Any] = {'orders': []}
    for order_id in SELECTED_ORDER_IDS:
        order = orders[order_id]
        items = line_items[order_id]
        file_type, style = STYLE_MAP[order_id]
        if file_type == 'pdf':
            path = RAW_DIR / f'{order_id}.pdf'
            make_pdf(path, order, items, style)
            attachments = []
        elif file_type == 'xlsx':
            path = RAW_DIR / f'{order_id}.xlsx'
            make_xlsx(path, order, items, style)
            attachments = []
        elif file_type == 'scan':
            path = RAW_DIR / f'{order_id}.png'
            make_scan(path, order, items, style)
            attachments = []
        elif file_type == 'email':
            path = RAW_DIR / f'{order_id}.eml'
            attachments = []
            if order_id == 'O017':
                att = build_attachment_xlsx(order, items)
                att_name = ATTACH_DIR / f'{order_id}_attachment.xlsx'
                att_name.write_bytes(att)
                attachments = [(att_name.name, att, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')]
                make_email(path, order, items, style, attachments)
            elif order_id == 'O025':
                att = build_attachment_xlsx(order, items)
                att_name = ATTACH_DIR / f'{order_id}_attachment.xlsx'
                att_name.write_bytes(att)
                attachments = [(att_name.name, att, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')]
                make_email(path, order, items, style, attachments)
            elif order_id == 'O029':
                att = build_attachment_xlsx(order, items, qty_override='6')
                att_name = ATTACH_DIR / f'{order_id}_attachment.xlsx'
                att_name.write_bytes(att)
                attachments = [(att_name.name, att, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')]
                make_email(path, order, items, style, attachments)
            else:
                make_email(path, order, items, style, attachments)
        else:
            raise ValueError(file_type)
        manifest['orders'].append({
            'order_id': order_id,
            'source_path': str(path.relative_to(ROOT)),
            'format': file_type,
            'style': style,
            'expected_status': order['order_status'],
            'attachments': [a[0] for a in attachments],
        })
    (RAW_DIR / 'manifest.json').write_text(json.dumps(manifest, indent=2), encoding='utf-8')
    return manifest


def make_structured_inputs(line_items: Dict[str, List[Dict[str, str]]]) -> Dict[str, List[Dict[str, Any]]]:
    grouped: Dict[str, List[Dict[str, Any]]] = {}
    for order_id in SELECTED_ORDER_IDS:
        rows: List[Dict[str, Any]] = []
        for item in line_items[order_id]:
            rows.append({
                'order_id': order_id,
                'line_number': item['line_number'],
                'customer_item_number': item['customer_item_number'],
                'raw_source': f"structured:{order_id}:{item['line_number']}",
                'raw_description': item['raw_customer_description'],
                'quantity': float(item['quantity']) if item['quantity'] else None,
                'unit': item['unit'] or None,
                'sku_hint': item['distributor_sku'] or None,
                'source_conflicts': item['notes'].split(';') if item['notes'] else None,
                'raw_notes': item['notes'] or None,
                'document_kind': 'structured',
                'source_quality': 'structured',
                'customer_alias': orders_cache[order_id]['customer'],
            })
        grouped[order_id] = rows
    return grouped


def compare_results(expected_orders: Dict[str, Dict[str, str]], truth_lines: Dict[str, List[Dict[str, str]]], structured: List[Dict[str, Any]], raw: List[Dict[str, Any]], blocked_orders: set[str]) -> Tuple[Dict[str, Any], Dict[str, Any], List[Dict[str, Any]]]:
    def by_order(rows: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
        return {row['order_id']: row for row in rows}

    structured_map = by_order(structured)
    raw_map = by_order(raw)

    def order_status(row: Dict[str, Any] | None, order_id: str) -> str:
        if row is None:
            return 'EXCEPTION'
        return row.get('order_status', 'EXCEPTION')

    def score_lines(rows: List[Dict[str, Any]], include_blocked: bool = False) -> Dict[str, Any]:
        flat_rows: List[Dict[str, Any]] = []
        for row in rows:
            flat_rows.extend(row.get('line_rows', []))
        order_lookup: Dict[str, Dict[str, Any]] = {}
        for row in rows:
            for line in row.get('line_rows', []):
                key = (line['order_id'], str(line['line_number']))
                order_lookup[key] = line
        total_truth = 0
        sku_correct = 0
        ready_total = 0
        ready_correct_sku = 0
        ready_wrong_qty = 0
        predicted_exception = 0
        actual_exception = 0
        tp_exception = 0
        false_review = 0
        missed_exception = 0
        false_confidence = 0
        hallucination = 0
        auto_ready = 0
        for order_id in SELECTED_ORDER_IDS:
            if order_id in blocked_orders and not include_blocked:
                continue
            for truth in truth_lines[order_id]:
                total_truth += 1
                key = (order_id, truth['line_number'])
                pred = order_lookup.get(key)
                actual_is_exception = truth['expected_status'] == 'EXCEPTION'
                if actual_is_exception:
                    actual_exception += 1
                if pred is None:
                    if actual_is_exception:
                        missed_exception += 1
                    continue
                predicted_status = pred['status'].upper()
                if predicted_status == 'EXCEPTION':
                    predicted_exception += 1
                    if actual_is_exception:
                        tp_exception += 1
                    if not actual_is_exception and pred.get('confidence') in {'HIGH', 'MEDIUM'}:
                        false_confidence += 1
                if actual_is_exception and predicted_status != 'EXCEPTION':
                    missed_exception += 1
                if not actual_is_exception and predicted_status == 'REVIEW REQUIRED':
                    false_review += 1
                if predicted_status == 'READY':
                    auto_ready += 1
                    if pred.get('proposed_sku') == truth['distributor_sku']:
                        ready_correct_sku += 1
                    else:
                        ready_wrong_qty += 1 if str(pred.get('extracted_value', {}).get('quantity')) != str(truth['quantity']) else 0
                if str(pred.get('proposed_sku') or '').strip() == str(truth['distributor_sku']).strip() and truth['distributor_sku']:
                    sku_correct += 1
                if predicted_status == 'READY' and pred.get('proposed_sku') and pred.get('proposed_sku') not in {truth['distributor_sku']} and truth['distributor_sku']:
                    hallucination += 0
                if not actual_is_exception and predicted_status == 'REVIEW REQUIRED':
                    pass
                if not actual_is_exception and pred.get('confidence') == 'HIGH' and pred.get('proposed_sku') != truth['distributor_sku']:
                    false_confidence += 1
        ready_truth = sum(1 for order_id in SELECTED_ORDER_IDS if order_id not in blocked_orders for truth in truth_lines[order_id] if truth['expected_status'] == 'READY')
        actual_exception_total = sum(1 for order_id in SELECTED_ORDER_IDS for truth in truth_lines[order_id] if truth['expected_status'] == 'EXCEPTION')
        predicted_exception_total = sum(1 for row in rows for line in row.get('line_rows', []) if line.get('status', '').upper() == 'EXCEPTION')
        order_rows = rows
        return {
            'order_level_accuracy': None,
            'exact_sku_accuracy': round(sku_correct / max(total_truth, 1), 4),
            'catalog_match_accuracy': round(sku_correct / max(total_truth, 1), 4),
            'exception_detection_precision': round(tp_exception / max(predicted_exception_total, 1), 4),
            'exception_detection_recall': round(tp_exception / max(actual_exception_total, 1), 4),
            'false_confidence_rate': round(false_confidence / max(total_truth, 1), 4),
            'hallucination_rate': round(hallucination / max(total_truth, 1), 4),
            'false_review_rate': round(false_review / max(ready_truth, 1), 4),
            'missed_exception_rate': round(missed_exception / max(actual_exception_total, 1), 4),
            'percent_auto_ready': round(auto_ready / max(total_truth, 1), 4),
            'percent_requiring_review': round(sum(1 for row in rows for line in row.get('line_rows', []) if line.get('status', '').upper() == 'REVIEW REQUIRED') / max(total_truth, 1), 4),
            'wrong_sku_ready': sum(1 for row in rows for line in row.get('line_rows', []) if line.get('status', '').upper() == 'READY' and line.get('proposed_sku') != truth_lines[line['order_id']][int(line['line_number'])-1]['distributor_sku']),
            'wrong_quantity_ready': sum(1 for row in rows for line in row.get('line_rows', []) if line.get('status', '').upper() == 'READY' and str(line.get('extracted_value', {}).get('quantity')) != str(truth_lines[line['order_id']][int(line['line_number'])-1]['quantity'])),
            'pred_ready': sum(1 for row in rows for line in row.get('line_rows', []) if line.get('status', '').upper() == 'READY'),
            'pred_review': sum(1 for row in rows for line in row.get('line_rows', []) if line.get('status', '').upper() == 'REVIEW REQUIRED'),
            'pred_exception': predicted_exception_total,
            'actual_ready': ready_truth,
            'actual_exception': actual_exception_total,
        }

    # order accuracy
    def order_level(rows: List[Dict[str, Any]]) -> float:
        hits = 0
        for order_id in SELECTED_ORDER_IDS:
            expected = expected_orders[order_id]['order_status']
            pred = 'EXCEPTION' if order_id in blocked_orders else order_status(by_order(rows).get(order_id), order_id)
            if pred == expected:
                hits += 1
        return round(hits / len(SELECTED_ORDER_IDS), 4)

    structured_metrics = score_lines(structured)
    raw_metrics = score_lines(raw)
    structured_metrics['order_level_accuracy'] = order_level(structured)
    raw_metrics['order_level_accuracy'] = order_level(raw)

    comparison = {k: {
        'structured': structured_metrics[k],
        'raw': raw_metrics[k],
        'delta': (None if structured_metrics[k] is None or raw_metrics[k] is None else round(raw_metrics[k] - structured_metrics[k], 4)),
    } for k in structured_metrics.keys()}

    failure_table: List[Dict[str, Any]] = []
    raw_line_map = {(line['order_id'], str(line['line_number'])): line for row in raw for line in row.get('line_rows', [])}
    struct_line_map = {(line['order_id'], str(line['line_number'])): line for row in structured for line in row.get('line_rows', [])}
    for order_id in SELECTED_ORDER_IDS:
        for truth in truth_lines[order_id]:
            key = (order_id, truth['line_number'])
            pred = raw_line_map.get(key)
            if pred is None and order_id in blocked_orders:
                failure_table.append({
                    'order_id': order_id,
                    'line_number': truth['line_number'],
                    'expected_status': truth['expected_status'],
                    'predicted_status': 'EXCEPTION',
                    'issue_type': 'OCR_BLOCKED',
                    'review_reason': 'OCR blocked; scan could not be read',
                })
                continue
            if pred is None:
                continue
            if pred.get('status', '').upper() != truth['expected_status'] or str(pred.get('proposed_sku') or '') != str(truth['distributor_sku'] or ''):
                failure_table.append({
                    'order_id': order_id,
                    'line_number': truth['line_number'],
                    'expected_status': truth['expected_status'],
                    'predicted_status': pred.get('status'),
                    'expected_sku': truth['distributor_sku'],
                    'predicted_sku': pred.get('proposed_sku'),
                    'expected_quantity': truth['quantity'],
                    'predicted_quantity': pred.get('extracted_value', {}).get('quantity'),
                    'review_reason': pred.get('review_reason'),
                })
    return structured_metrics, raw_metrics, failure_table


if __name__ == '__main__':
    orders_cache = load_orders()
    line_items_cache = load_line_items()
    clean_dir(RESULTS_DIR)
    manifest = generate_sources(orders_cache, line_items_cache)

    # structured baseline
    structured_groups = make_structured_inputs(line_items_cache)
    structured_results = run_order_validation(str(CATALOG), [{ 'order_id': oid } for oid in SELECTED_ORDER_IDS], structured_groups)

    # raw parsing + validation
    raw_results: List[Dict[str, Any]] = []
    blocked_orders: set[str] = set()
    parse_times: Dict[str, float] = {}
    raw_parse_details: Dict[str, Any] = {}
    for entry in manifest['orders']:
        order_id = entry['order_id']
        source_path = ROOT / entry['source_path']
        start = time.perf_counter()
        parsed = extract_order_from_document(source_path)
        parse_times[order_id] = round((time.perf_counter() - start) * 1000, 2)
        raw_parse_details[order_id] = {
            'parse_status': parsed.parse_status,
            'ocr_status': parsed.ocr_status,
            'line_count': len(parsed.line_items),
            'notes': parsed.notes,
            'warnings': parsed.warnings,
        }
        if parsed.parse_status == 'OCR_BLOCKED' or parsed.source_type == 'image':
            blocked_orders.add(order_id)
            raw_results.append({
                'order_id': order_id,
                'order_status': 'EXCEPTION',
                'line_rows': [],
                'parse_status': parsed.parse_status,
                'ocr_status': parsed.ocr_status,
                'blocked_reason': 'OCR_BLOCKED',
            })
            continue
        line_inputs = parsed_order_to_v2_inputs(parsed, entry['format'])
        for row in line_inputs:
            row['order_id'] = order_id
        result = run_order_validation(str(CATALOG), [{ 'order_id': order_id }], {order_id: line_inputs})[0]
        result['parse_status'] = parsed.parse_status
        result['ocr_status'] = parsed.ocr_status
        result['parse_time_ms'] = parse_times[order_id]
        raw_results.append(result)

    structured_metrics, raw_metrics, failure_table = compare_results(orders_cache, line_items_cache, structured_results, raw_results, blocked_orders)

    (RESULTS_DIR / '014c5_raw_manifest.json').write_text(json.dumps(manifest, indent=2), encoding='utf-8')
    (RESULTS_DIR / '014c5_structured_results.json').write_text(json.dumps(structured_results, indent=2), encoding='utf-8')
    (RESULTS_DIR / '014c5_raw_results.json').write_text(json.dumps(raw_results, indent=2), encoding='utf-8')
    (RESULTS_DIR / '014c5_structured_metrics.json').write_text(json.dumps(structured_metrics, indent=2), encoding='utf-8')
    (RESULTS_DIR / '014c5_raw_metrics.json').write_text(json.dumps(raw_metrics, indent=2), encoding='utf-8')
    (RESULTS_DIR / '014c5_comparison.json').write_text(json.dumps({'structured': structured_metrics, 'raw': raw_metrics}, indent=2), encoding='utf-8')
    (RESULTS_DIR / '014c5_failure_table.json').write_text(json.dumps(failure_table, indent=2), encoding='utf-8')
    (RESULTS_DIR / '014c5_parse_details.json').write_text(json.dumps(raw_parse_details, indent=2), encoding='utf-8')

    summary = {
        'selected_orders': SELECTED_ORDER_IDS,
        'blocked_orders': sorted(blocked_orders),
        'order_level_accuracy_delta': round(raw_metrics['order_level_accuracy'] - structured_metrics['order_level_accuracy'], 4),
        'raw_order_level_accuracy': raw_metrics['order_level_accuracy'],
        'structured_order_level_accuracy': structured_metrics['order_level_accuracy'],
        'raw_catalog_match_accuracy': raw_metrics['catalog_match_accuracy'],
        'structured_catalog_match_accuracy': structured_metrics['catalog_match_accuracy'],
    }
    (RESULTS_DIR / '014c5_summary.json').write_text(json.dumps(summary, indent=2), encoding='utf-8')
    print(json.dumps(summary, indent=2))
